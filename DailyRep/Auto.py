import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell

import win32com.client as win32
import os
import time


class Color():
    """Named colors for use in styles."""
    BLACK = 'FF000000'
    WHITE = 'FFFFFFFF'
    RED = 'FFFF0000'
    DARKRED = 'FF800000'
    BLUE = 'FF0000FF'
    DARKBLUE = 'FF000080'
    GREEN = 'FF00FF00'
    DARKGREEN = 'FF008000'
    YELLOW = 'FFFFFF00'
    DARKYELLOW = 'FF808000'


canada = "Canada"

UK_Sites = {"Agent": "B11",
            "Site": "B12",
            "UAR": "B13",
            "NULL": "B14",
            ">=95": "B4",
            ">94-95": "C4",
            ">92.5-94": "D4",
            ">91-92.5": "E4",
            "<91": "F4",
            "agentImpact": "B17",
            "siteImpact": "B18",
            "uarImpact": "B19",
            "assignment1": "A23",
            "assignment2": "A24",
            "assignment3": "A25",
            "Owner": "Prabir"
            }

ANZ_Sites = {"Agent": "C11",
             "Site": "C12",
             "UAR": "C13",
             "NULL": "C14",
             ">=95": "B5",
             ">94-95": "C5",
             ">92.5-94": "D5",
             ">91-92.5": "E5",
             "<91": "F5",
             "agentImpact": "C17",
             "siteImpact": "C18",
             "uarImpact": "C19",
             "assignment1": "A27",
             "assignment2": "A28",
             "assignment3": "A29",
             "Owner": "Nanditha"
             }

Indian_Sites = {"Agent": "D11",
                "Site": "D12",
                "UAR": "D13",
                "NULL": "D14",
                ">=95": "B6",
                ">94-95": "C6",
                ">92.5-94": "D6",
                ">91-92.5": "E6",
                "<91": "F6",
                "agentImpact": "D17",
                "siteImpact": "D18",
                "uarImpact": "D19",
                "assignment1": "A31",
                "assignment2": "A32",
                "assignment3": "A33",
                "Owner": "Abinash"
                }

SouthAfrican_Sites = {"Agent": "E11",
                      "Site": "E12",
                      "UAR": "E13",
                      "NULL": "E14",
                      ">=95": "B7",
                      ">94-95": "C7",
                      ">92.5-94": "D7",
                      ">91-92.5": "E7",
                      "<91": "F7",
                      "agentImpact": "E17",
                      "siteImpact": "E18",
                      "uarImpact": "E19",
                      "assignment1": "A35",
                      "assignment2": "A36",
                      "assignment3": "A37",
                      "Owner": "Vandana"
                      }

Canada = {"Agent": "F11",
          "Site": "F12",
          "UAR": "F13",
          "NULL": "F14",
          ">=95": "B8",
          ">94-95": "C8",
          ">92.5-94": "D8",
          ">91-92.5": "E8",
          "<91": "F8",
          "agentImpact": "F17",
          "siteImpact": "F18",
          "uarImpact": "F19",
          "assignment1": "A39",
          "assignment2": "A40",
          "assignment3": "A41",
          "Owner": "Amitabh"
          }

International = {
    ">=95": "B9",
    ">94-95": "C9",
    ">92.5-94": "D9",
    ">91-92.5": "E9",
    "<91": "F9",
}

greenFill = PatternFill(start_color='FF00FF00',
                        end_color='FF00FF00',
                        fill_type='solid')
yellowFill = PatternFill(start_color='FFFFFF00',
                         end_color='FFFFFF00',
                         fill_type='solid')
redFill = PatternFill(start_color='FFFF0000',
                      end_color='FFFF0000',
                      fill_type='solid')

blueFill = PatternFill(start_color='FF1e90FF',
                       end_color='FF1e90FF',
                       fill_type='solid')


def getDatafromSheet(wb, sheetName):
    caSheet = wb.get_sheet_by_name(sheetName)
    caAgentErr = caSheet['P5'].value
    caAgentErr = str(round(float(caAgentErr), 2))
    caSiteErr = caSheet['P6'].value
    caSiteErr = str(round(float(caSiteErr), 2))
    caUARErr = caSheet['P7'].value
    caUARErr = str(round(float(caUARErr), 2))
    caSuccess = caSheet['P9'].value
    caSuccess = str(round(float(caSuccess), 2))
    NUICIIErr = caSheet['P10'].value
    NUICIIErr = str(round(float(NUICIIErr), 2))

    data = {}
    data["agentError"] = caAgentErr
    data["siteError"] = caSiteErr
    data["uarError"] = caUARErr
    data["nullError"] = NUICIIErr
    data["success"] = round(float(caSuccess), 2)

    print sheetName, data

    return data


def fillData(wb, sheetName):
    data = getDatafromSheet(wb, sheetName)
    mainSheet = wb.get_sheet_by_name("Sheet1")
    mainSheet["A1"] = "Success Rating Details:"

    if sheetName == "UK_Sites":
        mainSheet[UK_Sites["Agent"]] = data["agentError"]
        mainSheet[UK_Sites["Site"]] = data["siteError"]
        mainSheet[UK_Sites["UAR"]] = data["uarError"]
        mainSheet[UK_Sites["NULL"]] = data["nullError"]

        if data["success"] >= 95:
            mainSheet[UK_Sites[">=95"]] = str(data["success"])
            mainSheet[UK_Sites[">=95"]].fill = blueFill
        if 94 <= data["success"] < 95:
            mainSheet[UK_Sites[">94-95"]] = str(data["success"])
            mainSheet[UK_Sites[">94-95"]].fill = greenFill
        if 92.5 <= data["success"] < 94:
            mainSheet[UK_Sites[">92.5-94"]] = str(data["success"])
            mainSheet[UK_Sites[">92.5-94"]].fill = yellowFill
        if 91 <= data["success"] < 92.5:
            mainSheet[UK_Sites[">91-92.5"]] = str(data["success"])
            mainSheet[UK_Sites[">91-92.5"]].fill = redFill
        if data["success"] < 91:
            mainSheet[UK_Sites["<91"]] = str(data["success"])
            mainSheet[UK_Sites["<91"]].fill = redFill

        wb.save("res.xlsx")
        return 0
    if sheetName == "Indian_Sites":
        mainSheet[Indian_Sites["Agent"]] = data["agentError"]
        mainSheet[Indian_Sites["Site"]] = data["siteError"]
        mainSheet[Indian_Sites["UAR"]] = data["uarError"]
        mainSheet[Indian_Sites["NULL"]] = data["nullError"]

        if data["success"] >= 95:
            mainSheet[Indian_Sites[">=95"]] = str(data["success"])
            mainSheet[Indian_Sites[">=95"]].fill = blueFill
        if 94 <= data["success"] < 95:
            mainSheet[Indian_Sites[">94-95"]] = str(data["success"])
            mainSheet[Indian_Sites[">94-95"]].fill = greenFill
        if 92.5 <= data["success"] < 94:
            mainSheet[Indian_Sites[">92.5-94"]] = str(data["success"])
            mainSheet[Indian_Sites[">92.5-94"]].fill = yellowFill
        if 91 <= data["success"] < 92.5:
            mainSheet[Indian_Sites[">91-92.5"]] = str(data["success"])
            mainSheet[Indian_Sites[">91-92.5"]].fill = redFill
        if data["success"] < 91:
            mainSheet[Indian_Sites["<91"]] = str(data["success"])
            mainSheet[Indian_Sites["<91"]].fill = redFill

        wb.save("res.xlsx")
        return 0
    if sheetName == "ANZ_Sites":
        mainSheet[ANZ_Sites["Agent"]] = data["agentError"]
        mainSheet[ANZ_Sites["Site"]] = data["siteError"]
        mainSheet[ANZ_Sites["UAR"]] = data["uarError"]
        mainSheet[ANZ_Sites["NULL"]] = data["nullError"]

        if data["success"] >= 95:
            mainSheet[ANZ_Sites[">=95"]] = str(data["success"])
            mainSheet[ANZ_Sites[">=95"]].fill = blueFill
        if 94 <= data["success"] < 95:
            mainSheet[ANZ_Sites[">94-95"]] = str(data["success"])
            mainSheet[ANZ_Sites[">94-95"]].fill = greenFill
        if 92.5 <= data["success"] < 94:
            mainSheet[ANZ_Sites[">92.5-94"]] = str(data["success"])
            mainSheet[ANZ_Sites[">92.5-94"]].fill = yellowFill
        if 91 <= data["success"] < 92.5:
            mainSheet[ANZ_Sites[">91-92.5"]] = str(data["success"])
            mainSheet[ANZ_Sites[">91-92.5"]].fill = redFill
        if data["success"] < 91:
            mainSheet[ANZ_Sites["<91"]] = str(data["success"])
            mainSheet[ANZ_Sites["<91"]].fill = redFill

        wb.save("res.xlsx")
        return 0
    if sheetName == "SouthAfrican_Sites":
        mainSheet[SouthAfrican_Sites["Agent"]] = data["agentError"]
        mainSheet[SouthAfrican_Sites["Site"]] = data["siteError"]
        mainSheet[SouthAfrican_Sites["UAR"]] = data["uarError"]
        mainSheet[SouthAfrican_Sites["NULL"]] = data["nullError"]

        if data["success"] >= 95:
            mainSheet[SouthAfrican_Sites[">=95"]] = str(data["success"])
            mainSheet[SouthAfrican_Sites[">=95"]].fill = blueFill
        if 94 <= data["success"] < 95:
            mainSheet[SouthAfrican_Sites[">94-95"]] = str(data["success"])
            mainSheet[SouthAfrican_Sites[">94-95"]].fill = greenFill
        if 92.5 <= data["success"] < 94:
            mainSheet[SouthAfrican_Sites[">92.5-94"]] = str(data["success"])
            mainSheet[SouthAfrican_Sites[">92.5-94"]].fill = yellowFill
        if 91 <= data["success"] < 92.5:
            mainSheet[SouthAfrican_Sites[">91-92.5"]] = str(data["success"])
            mainSheet[SouthAfrican_Sites[">91-92.5"]].fill = redFill
        if data["success"] < 91:
            mainSheet[SouthAfrican_Sites["<91"]] = str(data["success"])
            mainSheet[SouthAfrican_Sites["<91"]].fill = redFill

        wb.save("res.xlsx")
        return 0
    if sheetName == "Canada":
        mainSheet[Canada["Agent"]] = data["agentError"]
        mainSheet[Canada["Site"]] = data["siteError"]
        mainSheet[Canada["UAR"]] = data["uarError"]
        mainSheet[Canada["NULL"]] = data["nullError"]

        if data["success"] >= 95:
            mainSheet[Canada[">=95"]] = str(data["success"])
            mainSheet[Canada[">=95"]].fill = blueFill
        if 94 <= data["success"] < 95:
            mainSheet[Canada[">94-95"]] = str(data["success"])
            mainSheet[Canada[">94-95"]].fill = greenFill
        if 92.5 <= data["success"] < 94:
            mainSheet[Canada[">92.5-94"]] = str(data["success"])
            mainSheet[Canada[">92.5-94"]].fill = yellowFill
        if 91 <= data["success"] < 92.5:
            mainSheet[Canada[">91-92.5"]] = str(data["success"])
            mainSheet[Canada[">91-92.5"]].fill = redFill
        if data["success"] < 91:
            mainSheet[Canada["<91"]] = str(data["success"])
            mainSheet[Canada["<91"]].fill = redFill

        wb.save("res.xlsx")
        return 0

        # wb.save("res.xlsx")


# fillData(wb, "Indian_Sites")
# fillData(wb, "SouthAfrican_Sites")
# fillData(wb, "Indian_Sites")
# fillData(wb, "Indian_Sites")
#
# def returnAgentNamefromCell(cell, sheetName):
#     cell = "A" + cell[:-1]
#
#     return mainSheet[cell].value

def changeToAssignment(index):
    return "B" + index[1:]


def returnAgentNamefromImpactDict(wb, dicOfImpactDict, sheetName):
    agentDict = dicOfImpactDict['agentImpact']
    siteDict = dicOfImpactDict['siteImpact']
    uarDict = dicOfImpactDict['uarImpact']

    print "debug 1"
    print dicOfImpactDict["agentImpact"]
    print dicOfImpactDict["siteImpact"]
    print dicOfImpactDict["uarImpact"]

    kIndexAgent = [i[0] for i in agentDict]
    kIndexSite = [i[0] for i in siteDict]
    kIndexIUAR = [i[0] for i in uarDict]

    print "debug 2"
    print kIndexAgent
    print kIndexSite
    print kIndexIUAR

    valIndexAgent = [str(round(float(i[1]), 2)) for i in agentDict]
    valIndexSite = [str(round(float(i[1]), 2)) for i in siteDict]
    valIndexUAR = [str(round(float(i[1]), 2)) for i in uarDict]

    print "debug 3"
    print valIndexAgent
    print valIndexSite
    print valIndexUAR

    # print kIndexAgent
    # print kIndexSite
    # print kIndexIUAR

    agentDict = ["A" + j[1:] for j in kIndexAgent]
    print "debugging"
    print agentDict
    siteDict = ["A" + j[1:] for j in kIndexSite]
    uarDict = ["A" + j[1:] for j in kIndexIUAR]

    print "debug"
    print agentDict
    print siteDict
    print uarDict

    mainSheet = wb.get_sheet_by_name(sheetName)

    agentImpactNames = [mainSheet[i].value for i in agentDict]
    print agentImpactNames
    agentImpactAssignment = agentImpactNames
    siteImpactNames = [mainSheet[i].value for i in siteDict]
    uarImpactNames = [mainSheet[i].value for i in uarDict]

    print "debug"
    print siteImpactNames
    print uarImpactNames

    # if sheetName == "UK_Sites":
    #     mainSheet[UK_Sites["assignment1"]] = agentImpactNames[0]
    #     mainSheet[UK_Sites["assignment2"]] = agentImpactNames[1]
    #     mainSheet[UK_Sites["assignment3"]] = agentImpactNames[2]
    #
    #     wb.save("res.xlsx")

    x = zip(agentImpactNames, valIndexAgent)
    y = zip(siteImpactNames, valIndexSite)
    z = zip(uarImpactNames, valIndexUAR)

    agentImpactNames = [i[0] + "(" + i[1] + "%)" for i in x]
    siteImpactNames = [i[0] + "(" + i[1] + "%)" for i in y]
    uarImpactNames = [i[0] + "(" + i[1] + "%)" for i in z]

    agentNameagent = ", \n".join(agentImpactNames)
    agentNamesite = ", \n".join(siteImpactNames)
    agentNameuar = ", \n".join(uarImpactNames)

    # wb.get_sheet_by_name("Sheet1").column_dimensions["B"].width = 30
    # wb.get_sheet_by_name("Sheet1").column_dimensions["C"].width = 30
    # wb.get_sheet_by_name("Sheet1").column_dimensions["D"].width = 30
    # wb.get_sheet_by_name("Sheet1").column_dimensions["E"].width = 30
    # wb.get_sheet_by_name("Sheet1").column_dimensions["F"].width = 30


    if sheetName == "UK_Sites":
        mainSheet = wb.get_sheet_by_name("Sheet1")
        # mainSheet.column_dimensions(UK_Sites["agentImpact"]).width = 27
        # mainSheet.column_dimensions[UK_Sites["siteImpact"]].width = 27
        # mainSheet.column_dimensions[UK_Sites["uarImpact"]].width = 27
        mainSheet.cell(UK_Sites["agentImpact"]).style.alignment.wrap_text = True
        mainSheet.cell(UK_Sites["siteImpact"]).style.alignment.wrap_text = True
        mainSheet.cell(UK_Sites["uarImpact"]).style.alignment.wrap_text = True
        mainSheet[UK_Sites["agentImpact"]] = agentNameagent
        mainSheet[UK_Sites["siteImpact"]] = agentNamesite
        mainSheet[UK_Sites["uarImpact"]] = agentNameuar
        # adding assignments
        mainSheet[UK_Sites["assignment1"]] = agentImpactAssignment[0]
        mainSheet[UK_Sites["assignment2"]] = agentImpactAssignment[1]
        mainSheet[UK_Sites["assignment3"]] = agentImpactAssignment[2]

        mainSheet[changeToAssignment(UK_Sites["assignment1"])] = UK_Sites["Owner"]
        mainSheet[changeToAssignment(UK_Sites["assignment2"])] = UK_Sites["Owner"]
        mainSheet[changeToAssignment(UK_Sites["assignment3"])] = UK_Sites["Owner"]
        # mainSheet[UK_Sites["assignment2"]] = agentImpactAssignment[1]
        # mainSheet[UK_Sites["assignment3"]] = agentImpactAssignment[2]

        wb.save("res.xlsx")
        return 0
    if sheetName == "Indian_Sites":
        mainSheet = wb.get_sheet_by_name("Sheet1")
        mainSheet.cell(Indian_Sites["agentImpact"]).style.alignment.wrap_text = True
        mainSheet.cell(Indian_Sites["siteImpact"]).style.alignment.wrap_text = True
        mainSheet.cell(Indian_Sites["uarImpact"]).style.alignment.wrap_text = True
        mainSheet[Indian_Sites["agentImpact"]] = agentNameagent
        mainSheet[Indian_Sites["siteImpact"]] = agentNamesite
        mainSheet[Indian_Sites["uarImpact"]] = agentNameuar
        # adding assignments
        mainSheet[Indian_Sites["assignment1"]] = agentImpactAssignment[0]
        mainSheet[Indian_Sites["assignment2"]] = agentImpactAssignment[1]
        mainSheet[Indian_Sites["assignment3"]] = agentImpactAssignment[2]

        mainSheet[changeToAssignment(Indian_Sites["assignment1"])] = Indian_Sites["Owner"]
        mainSheet[changeToAssignment(Indian_Sites["assignment2"])] = Indian_Sites["Owner"]
        mainSheet[changeToAssignment(Indian_Sites["assignment3"])] = Indian_Sites["Owner"]

        wb.save("res.xlsx")
        return 0
    if sheetName == "ANZ_Sites":
        mainSheet = wb.get_sheet_by_name("Sheet1")
        mainSheet.cell(UK_Sites["agentImpact"]).style.alignment.wrap_text = True
        mainSheet.cell(UK_Sites["siteImpact"]).style.alignment.wrap_text = True
        mainSheet.cell(UK_Sites["uarImpact"]).style.alignment.wrap_text = True
        mainSheet[ANZ_Sites["agentImpact"]] = agentNameagent
        mainSheet[ANZ_Sites["siteImpact"]] = agentNamesite
        mainSheet[ANZ_Sites["uarImpact"]] = agentNameuar
        # adding assignments
        mainSheet[ANZ_Sites["assignment1"]] = agentImpactAssignment[0]
        mainSheet[ANZ_Sites["assignment2"]] = agentImpactAssignment[1]
        mainSheet[ANZ_Sites["assignment3"]] = agentImpactAssignment[2]

        mainSheet[changeToAssignment(ANZ_Sites["assignment1"])] = ANZ_Sites["Owner"]
        mainSheet[changeToAssignment(ANZ_Sites["assignment2"])] = ANZ_Sites["Owner"]
        mainSheet[changeToAssignment(ANZ_Sites["assignment3"])] = ANZ_Sites["Owner"]

        wb.save("res.xlsx")
        return 0

    if sheetName == "SouthAfrican_Sites":
        mainSheet = wb.get_sheet_by_name("Sheet1")
        mainSheet.cell(ANZ_Sites["agentImpact"]).style.alignment.wrap_text = True
        mainSheet.cell(ANZ_Sites["siteImpact"]).style.alignment.wrap_text = True
        mainSheet.cell(ANZ_Sites["uarImpact"]).style.alignment.wrap_text = True
        mainSheet[SouthAfrican_Sites["agentImpact"]] = agentNameagent
        mainSheet[SouthAfrican_Sites["siteImpact"]] = agentNamesite
        mainSheet[SouthAfrican_Sites["uarImpact"]] = agentNameuar
        # adding assignments
        mainSheet[SouthAfrican_Sites["assignment1"]] = agentImpactAssignment[0]
        mainSheet[SouthAfrican_Sites["assignment2"]] = agentImpactAssignment[1]
        mainSheet[SouthAfrican_Sites["assignment3"]] = agentImpactAssignment[2]

        mainSheet[changeToAssignment(SouthAfrican_Sites["assignment1"])] = SouthAfrican_Sites["Owner"]
        mainSheet[changeToAssignment(SouthAfrican_Sites["assignment2"])] = SouthAfrican_Sites["Owner"]
        mainSheet[changeToAssignment(SouthAfrican_Sites["assignment3"])] = SouthAfrican_Sites["Owner"]

        wb.save("res.xlsx")
        return 0

    if sheetName == "Canada":
        mainSheet = wb.get_sheet_by_name("Sheet1")
        mainSheet.cell(Canada["agentImpact"]).style.alignment.wrap_text = True
        mainSheet.cell(Canada["siteImpact"]).style.alignment.wrap_text = True
        mainSheet.cell(Canada["uarImpact"]).style.alignment.wrap_text = True
        mainSheet[Canada["agentImpact"]] = agentNameagent
        mainSheet[Canada["siteImpact"]] = agentNamesite
        mainSheet[Canada["uarImpact"]] = agentNameuar

        # adding assignments
        mainSheet[Canada["assignment1"]] = agentImpactAssignment[0]
        mainSheet[Canada["assignment2"]] = agentImpactAssignment[1]
        mainSheet[Canada["assignment3"]] = agentImpactAssignment[2]

        mainSheet[changeToAssignment(Canada["assignment1"])] = Canada["Owner"]
        mainSheet[changeToAssignment(Canada["assignment2"])] = Canada["Owner"]
        mainSheet[changeToAssignment(Canada["assignment3"])] = Canada["Owner"]

        wb.save("res.xlsx")
        return 0



        # print zip(kIndexAgent, agentImpactNames)
        # print zip(kIndexSite, siteImpactNames)
        # print zip(kIndexIUAR, uarImpactNames)


        # print agentDict
        # print siteDict
        # print uarDict


def fillFormula(wb, sheetName):
    mainSheet = wb.get_sheet_by_name(sheetName)
    mainSheet["K1"] = "Agent Impact"
    mainSheet["L1"] = "Site Impact"
    mainSheet["M1"] = "UAR Impact"

    # Getting value of total cell of the sheet
    total = "B" + str(mainSheet.max_row)
    t = float(mainSheet[total].value)

    agentImpactDict = {}
    siteImpactDict = {}
    uarImpactDict = {}

    for i in range(2, mainSheet.max_row - 1):
        i = str(i)

        # Agent impact

        formCell = "K" + i
        agentErrCell = "D" + i
        # totalCell = total

        # calc agent impact for dict
        agentImpact = float(mainSheet[agentErrCell].value) / t * 100

        mainSheet[formCell] = "=" + agentErrCell + "/" + str(t) + "*100"

        agentImpactDict[formCell] = agentImpact


        # Site impact

        formCell = "L" + i
        siteErrCell = "E" + i

        siteImpact = float(mainSheet[siteErrCell].value) / t * 100

        mainSheet[formCell] = "=" + siteErrCell + "/" + str(t) + "*100"

        siteImpactDict[formCell] = siteImpact

        # UAR impact

        formCell = "M" + i
        uarErrCell = "F" + i

        uarImpact = float(mainSheet[uarErrCell].value) / t * 100

        mainSheet[formCell] = "=" + uarErrCell + "/" + str(t) + "*100"

        uarImpactDict[formCell] = uarImpact

    agentImpactDict = sorted(agentImpactDict.items(), key=lambda x: x[1], reverse=True)
    siteImpactDict = sorted(siteImpactDict.items(), key=lambda x: x[1], reverse=True)
    uarImpactDict = sorted(uarImpactDict.items(), key=lambda x: x[1], reverse=True)

    dicOfImpactDict = {}
    dicOfImpactDict["agentImpact"] = agentImpactDict[:3]
    dicOfImpactDict["siteImpact"] = siteImpactDict[:3]
    dicOfImpactDict["uarImpact"] = uarImpactDict[:3]

    print "debug dicOfImpactDict"
    print dicOfImpactDict

    returnAgentNamefromImpactDict(wb, dicOfImpactDict, sheetName)

    # print agentImpactDict
    # print siteImpactDict
    # print uarImpactDict
    wb.save("res.xlsx")


def calcInternational(wb):
    totalSuccess = []
    total = []
    sheets = wb.get_sheet_names()
    sheets.remove("ICICI_Sites")
    sheets.remove("Ireland")
    sheets.remove("RBC")
    sheets.remove("Sheet1")
    for i in sheets:
        mainSheet = wb.get_sheet_by_name(i)
        indexTotal = "B" + str(mainSheet.max_row)
        indexSuccess = "C" + str(mainSheet.max_row)
        total.append(float(mainSheet[indexTotal].value))
        totalSuccess.append(float(mainSheet[indexSuccess].value))
    print sum(totalSuccess) / sum(total) * 100

    data = {}
    internationalSuccess = sum(totalSuccess) / sum(total) * 100
    internationalSuccess = round(float(internationalSuccess), 2)
    data["success"] = internationalSuccess
    mainSheet = wb.get_sheet_by_name("Sheet1")

    if data["success"] >= 95:
        mainSheet[International[">=95"]] = str(data["success"])
        mainSheet[International[">=95"]].fill = blueFill
    if 94 <= data["success"] < 95:
        mainSheet[International[">94-95"]] = str(data["success"])
        mainSheet[International[">94-95"]].fill = greenFill
    if 92.5 <= data["success"] < 94:
        mainSheet[International[">92.5-94"]] = str(data["success"])
        mainSheet[International[">92.5-94"]].fill = yellowFill
    if 91 <= data["success"] < 92.5:
        mainSheet[International[">91-92.5"]] = str(data["success"])
        mainSheet[International[">91-92.5"]].fill = redFill
    if data["success"] < 91:
        mainSheet[International["<91"]] = str(data["success"])
        mainSheet[International["<91"]].fill = redFill

    wb.save("res.xlsx")


def populateDataAuto():
    wb = openpyxl.load_workbook('stats.xlsx', data_only=True)
    sheets = wb.get_sheet_names()
    print sheets

    sheets.remove("ICICI_Sites")
    sheets.remove("Ireland")
    sheets.remove("RBC")
    sheets.remove("Sheet1")

    calcInternational(wb)
    x = ['Canada', 'Indian_Sites', 'ANZ_Sites', 'UK_Sites', 'SouthAfrican_Sites']
    for i in sheets:
        if i not in x:
            continue
        fillData(wb, i)
        fillFormula(wb, i)

    print "Stats Added"

# populateDataAuto()
# getAgentImpact(wb, i)
# fillFormula(wb, 31)
# fillFormula(wb, "Canada")
