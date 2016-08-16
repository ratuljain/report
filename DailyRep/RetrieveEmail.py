# -*- coding: utf-8 -*-
"""
Outlook folder email reader

Created on   : 25/09/2015
"""
__author__ = 'ivanlla'

import win32com.client
import win32com.client as win32
import os
import time
import openpyxl

import time
# dd/mm/yyyy format
todayDate = time.strftime("%m/%d/%Y")
searchString = "I18N Locales report for TTR Agents " + todayDate
# print searchString

try:
    os.remove("stats.xls")
    os.remove("stats.xlsx")
except OSError:
    pass

DEBUG = False
EXCLUSION_LIST = [
    'SharePoint Lists - outlook',
    'Project Canendars'
]

outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")


class Oli(object):
    def __init__(self, outlook_object):
        self._obj = outlook_object

    def items(self):
        array_size = self._obj.Count
        for item_index in xrange(1, array_size + 1):
            yield (item_index, self._obj[item_index - 1])

    def prop(self):
        return sorted(self._obj._prop_map_get_.keys())


def search_item(folders, name):
    if DEBUG: browse(folders, recursive=False)
    for index, folder in Oli(folders).items():
        if folder.Name == name:
            if DEBUG: print " Found %s @ %d" % (folder.Name, index)
            return index, folder
    return None, None


def search(path):
    components = path.split('/')
    if DEBUG: print components
    folder = None
    root = outlook.folders
    for name in components:
        index, folder = search_item(root, name)
        if not index:
            return None
        root = folder.Folders

    return folder


def browse(folders, depth=2, recursive=True):
    if not folders:
        return
    for index, folder in Oli(folders).items():
        print " " * depth, u"(%i) [%s] [%s]" % (index, folder.Name, folder)
        if u"%s" % folder in EXCLUSION_LIST:
            continue
        if recursive:
            browse(folder.Folders, depth + 2, recursive)


def process_messages(folder):
    messageList = []
    if not folder:
        print "Folder could not be found!"
        return
    messages = folder.Items
    message = messages.GetFirst()
    while message:
        # Process a message
        if "I18N Locales report for TTR Agents" in message.Subject and todayDate in message.Subject:
            messageList.append(message)
            print "yussss"
            print "%s;%s" % (message.Subject, message.CreationTime)
        message = messages.GetNext()

    return messageList


def getMail():
    # if __name__ == "__main__":
    # list(outlook.Folders)
    f = outlook.GetDefaultFolder(6)
    if DEBUG and f: print "Folder name: ", f.Name
    msgList = process_messages(f)
    message = msgList[-1]

    attachments = message.Attachments
    attachment = attachments.Item(1)
    attachment.SaveAsFile(os.getcwd() + '\\stats.xls')
    print os.getcwd() + '\\stats.xls'
    time.sleep(10)
    print "Mail saved"


#

def convertToXlsx():
    print os.getcwd()
    fname = os.getcwd() + '\\stats.xls'
    saveLocation = ""
    print fname
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(fname)

    wb.SaveAs(fname + "x", FileFormat=51)  # FileFormat = 51 is for .xlsx extension
    wb.Close()  # FileFormat = 56 is for .xls extension
    excel.Application.Quit()
    print "Converted to XLSX"

    time.sleep(3)

    wb = openpyxl.load_workbook('stats.xlsx', data_only=True)
    mainSheet = wb.get_sheet_by_name("Indian_Sites")
    mainSheet["B" + str(mainSheet.max_row)] = str(int(mainSheet["B" + str(mainSheet.max_row)].value) +
                                                  int(mainSheet["H" + str(mainSheet.max_row)].value) +
                                                  int(mainSheet["I" + str(mainSheet.max_row)].value) +
                                                  int(mainSheet["J" + str(mainSheet.max_row)].value))

    mainSheet["D" + str(mainSheet.max_row)] = str(int(mainSheet["D" + str(mainSheet.max_row)].value) +
                                                  int(mainSheet["G" + str(mainSheet.max_row)].value))

    print mainSheet["D" + str(mainSheet.max_row)].value

    mainSheet["E" + str(mainSheet.max_row)] = str(int(mainSheet["E" + str(mainSheet.max_row)].value) +
                                                  int(mainSheet["I" + str(mainSheet.max_row)].value) +
                                                  int(mainSheet["J" + str(mainSheet.max_row)].value))

    print mainSheet["E" + str(mainSheet.max_row)].value

    mainSheet["F" + str(mainSheet.max_row)] = str(int(mainSheet["F" + str(mainSheet.max_row)].value) +
                                                  int(mainSheet["H" + str(mainSheet.max_row)].value))

    print mainSheet["F" + str(mainSheet.max_row)].value

    wb.save("stats.xlsx")

    print "data changed"
