__author__ = 'rjain1'
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
import string

cells = ['A2', 'A3', 'A4']

UK_Sites = {"Agent": "B11",
            "Site": "B12",
            "UAR": "B13",
            "NULL": "B14",
            ">=95": "B4",
            ">94-95": "C4",
            ">92.5-94": "D4",
            ">91-92.5": "E4",
            "<91": "F4",
            "agentImpact": "B17",
            "siteImpact": "B18",
            "uarImpact": "B19",
            "assignment1": "A23",
            "assignment2": "A24",
            "assignment3": "A25",
            "Owner": "Shefali"
            }

ANZ_Sites = {"Agent": "C11",
             "Site": "C12",
             "UAR": "C13",
             "NULL": "C14",
             ">=95": "B5",
             ">94-95": "C5",
             ">92.5-94": "D5",
             ">91-92.5": "E5",
             "<91": "F5",
             "agentImpact": "C17",
             "siteImpact": "C18",
             "uarImpact": "C19",
             "assignment1": "A27",
             "assignment2": "A28",
             "assignment3": "A29",
             "Owner": "Vivek"
             }

Indian_Sites = {"Agent": "D11",
                "Site": "D12",
                "UAR": "D13",
                "NULL": "D14",
                ">=95": "B6",
                ">94-95": "C6",
                ">92.5-94": "D6",
                ">91-92.5": "E6",
                "<91": "F6",
                "agentImpact": "D17",
                "siteImpact": "D18",
                "uarImpact": "D19",
                "assignment1": "A31",
                "assignment2": "A32",
                "assignment3": "A33",
                "Owner": "Abinash"
                }

SouthAfrican_Sites = {"Agent": "E11",
                      "Site": "E12",
                      "UAR": "E13",
                      "NULL": "E14",
                      ">=95": "B7",
                      ">94-95": "C7",
                      ">92.5-94": "D7",
                      ">91-92.5": "E7",
                      "<91": "F7",
                      "agentImpact": "E17",
                      "siteImpact": "E18",
                      "uarImpact": "E19",
                      "assignment1": "A35",
                      "assignment2": "A36",
                      "assignment3": "A37",
                      "Owner": "Nanditha"
                      }

Canada = {"Agent": "F11",
          "Site": "F12",
          "UAR": "F13",
          "NULL": "F14",
          ">=95": "B8",
          ">94-95": "C8",
          ">92.5-94": "D8",
          ">91-92.5": "E8",
          "<91": "F8",
          "agentImpact": "F17",
          "siteImpact": "F18",
          "uarImpact": "F19",
          "assignment1": "A39",
          "assignment2": "A40",
          "assignment3": "A41",
          "Owner": "Ratul"
          }

International = {
    ">=95": "B9",
    ">94-95": "C9",
    ">92.5-94": "D9",
    ">91-92.5": "E9",
    "<91": "F9",
}


# print cells

# ws = wb.get_sheet_by_name("UK_Sites")
# allRows = list(ws.iter_rows())
#
# res = [i[:13] for i in allRows if i[0].row in cells]
# print res
# for i in allRows:
#     print i[0].row

def writeRowfromList(wb, data, index):
    mainSheet = wb.get_sheet_by_name("Sheet1")
    print data

    columns = len(data)
    alpha = list(string.uppercase)[:columns]
    sourceIndex = [alpha[i] + str(index) for i in range(len(alpha))]

    for i in xrange(len(sourceIndex)):
        mainSheet[sourceIndex[i]] = data[i]

    wb.save("res.xlsx")


def writeAssignmentStats(wb, data, sheetName):
    i = 0

    if sheetName == "UK_Sites":
        startIndex = 23
        endIndex = 25
    if sheetName == "ANZ_Sites":
        startIndex = 27
        endIndex = 29
    if sheetName == "Indian_Sites":
        startIndex = 31
        endIndex = 33
    if sheetName == "SouthAfrican_Sites":
        startIndex = 35
        endIndex = 37
    if sheetName == "Canada":
        startIndex = 39
        endIndex = 41

    for index in xrange(startIndex, endIndex + 1):
        if i > 2:
            break
        writeRowfromList(wb, data[i], index)
        i += 1


def getRowsStats(wb, sheetName, cells):
    res = []
    cells = [int(i[1]) for i in cells]
    g = lambda x: [i.value for i in x]
    ws = wb.get_sheet_by_name(sheetName)

    allRows = list(ws.iter_rows())
    res = [i[:6] for i in allRows if i[0].row in cells]
    rest = [g(i) for i in res]
    # print rest
    return rest


def doAssignments():
    wb = openpyxl.load_workbook('res.xlsx', data_only=True)

    locales = ['Canada', 'Indian_Sites', 'ANZ_Sites', 'UK_Sites', 'SouthAfrican_Sites']

    for i in locales:
        data = getRowsStats(wb, i, cells)
        writeAssignmentStats(wb, data, i)
        # writeAssignmentStats(wb, x)
