__author__ = 'rjain1'
from openpyxl.reader.excel import load_workbook
import itertools
import string
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment, Side
from win32com.client import Dispatch
import os, time

ratingRows = ['A2', 'B2', 'C2', 'D2', 'E2', 'F2', 'A22', 'B22']
localeRows = ['A10', 'B10', 'C10', 'D10', 'E10', 'F10', 'G10', 'A16', 'B16', 'C16', 'D16', 'E16', 'F16']
assignmentRows = [i + '22' for i in list(string.uppercase)[:6]]
targetRows = ['A3', 'B3', 'C3', 'D3', 'E3', 'F3']
grayColumn = ["A" + str(i) for i in range(4, 20) if i not in [15]]


def createSingleRow(indexAlpha, num):
    indexAlpha = [i + str(num) for i in indexAlpha]
    return indexAlpha


alpha = ['B', 'C', 'D', 'E', 'F', 'G']
temp = [createSingleRow(alpha, i) for i in range(4, 20)]

wrapRows = list(itertools.chain(*temp))
rem = ["G" + str(i) for i in range(1, 16)]
allCellstemp = [createSingleRow(alpha + ['A'], i) for i in range(1, 42)]
allCells = list(itertools.chain(*allCellstemp))
#
# blueFill = PatternFill(start_color='FF1e90FF',
#                        end_color='FF1e90FF',
#                        fill_type='solid')

grayFill = PatternFill(start_color='FFD9D9D9',
                       end_color='FFD9D9D9',
                       fill_type='solid')

ft = Font(name='Cambria',
          size=10,
          bold=True,
          italic=True,
          vertAlign=None,
          underline='none',
          strike=False,
          color='FFFFFFFF')

ft3 = Font(name='Cambria',
           size=10,
           bold=True,
           italic=True,
           vertAlign=None,
           underline='none',
           strike=False,
           color='FFFFD966')

ft1 = Font(name='Cambria',
           size=10,
           bold=True,
           italic=False,
           vertAlign=None,
           underline='none',
           strike=False,
           color='FF000000')

ft2 = Font(name='Cambria',
           size=10,
           bold=False,
           italic=False,
           vertAlign=None,
           underline='none',
           strike=False,
           color='FF000000')

alignment = Alignment(horizontal='general',
                      vertical='bottom',
                      text_rotation=0,
                      wrap_text=True,
                      shrink_to_fit=True,
                      indent=0)

alignmentR = Alignment(horizontal='right',
                       vertical='bottom',
                       text_rotation=0,
                       wrap_text=True,
                       shrink_to_fit=True,
                       indent=0)

alignmentL = Alignment(horizontal='left',
                       vertical='bottom',
                       text_rotation=0,
                       wrap_text=True,
                       shrink_to_fit=True,
                       indent=0)

border = Border(left=Side(border_style=None,
                          color='FF000000'),
                right=Side(border_style=None,
                           color='FF000000'),
                top=Side(border_style=None,
                         color='FF000000'),
                bottom=Side(border_style=None,
                            color='FF000000'),
                diagonal=Side(border_style=None,
                              color='FF000000'),
                diagonal_direction=0,
                outline=Side(border_style=None,
                             color='FF000000'),
                vertical=Side(border_style=None,
                              color='FF000000'),
                horizontal=Side(border_style=None,
                                color='FF000000')
                )

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

blueFill = PatternFill(start_color='FF1F4E78',
                       end_color='FF1F4E78',
                       fill_type='solid')


def formatRatingRow(wb, ratingRows, localeRows, assignmentRows):
    mainSheet = wb.get_sheet_by_name("Sheet1")

    for i in ratingRows + localeRows + assignmentRows:
        cell = mainSheet[i]
        cell.font = ft
        cell.fill = blueFill

    for i in assignmentRows:
        cell = mainSheet[i]
        cell.font = ft3

    mainSheet['G22'] = "Assignment To"
    mainSheet['G22'].border = thin_border
    mainSheet['G22'].font = ft3
    mainSheet['G22'].fill = blueFill

        # wb.save("stats.xlsx")


def formatGrayColumn(wb, targetRows, grayColumn):
    mainSheet = wb.get_sheet_by_name("Sheet1")

    for i in targetRows + grayColumn:
        cell = mainSheet[i]
        cell.font = ft1
        cell.fill = grayFill

        # wb.save("stats.xlsx")


def changeDim(wb, wrapRows):
    mainSheet = wb.get_sheet_by_name("Sheet1")

    wb.get_sheet_by_name("Sheet1").column_dimensions["A"].width = 20
    wb.get_sheet_by_name("Sheet1").column_dimensions["B"].width = 38
    wb.get_sheet_by_name("Sheet1").row_dimensions[17].height = 48
    wb.get_sheet_by_name("Sheet1").row_dimensions[18].height = 48
    wb.get_sheet_by_name("Sheet1").row_dimensions[19].height = 48
    wb.get_sheet_by_name("Sheet1").column_dimensions["C"].width = 38
    wb.get_sheet_by_name("Sheet1").column_dimensions["D"].width = 38
    wb.get_sheet_by_name("Sheet1").column_dimensions["E"].width = 38
    wb.get_sheet_by_name("Sheet1").column_dimensions["F"].width = 38
    wb.get_sheet_by_name("Sheet1").column_dimensions["G"].width = 38
    wb.get_sheet_by_name("Sheet1").column_dimensions["H"].width = 38
    wb.get_sheet_by_name("Sheet1").column_dimensions["I"].width = 38
    wb.get_sheet_by_name("Sheet1").column_dimensions["j"].width = 38

    for i in wrapRows:
        cell = mainSheet[i]
        cell.alignment = alignment
    cell = mainSheet['A1']
    cell.alignment = alignment


    # wb.save("stats.xlsx")


def set_border(wb, cell_range):
    mainSheet = wb.get_sheet_by_name("Sheet1")


    for i in cell_range:
        cell = mainSheet[i]
        cell.border = thin_border
        cell.font = ft2


def autoFit():
    excel = Dispatch('Excel.Application')
    wb = excel.Workbooks.Open(os.getcwd() + '\\res.xlsx')

    # Activate second sheet
    excel.Worksheets(9).Activate()

    # Autofit column in active sheet
    excel.ActiveSheet.Columns.AutoFit()
    excel.ActiveSheet.Rows.AutoFit()

    # Save changes in a new file
    # wb.SaveAs(os.getcwd() + '\\res1.xlsx')

    # Or simply save changes in a current file
    wb.Save()

    wb.Close()
    print "Autofit done"



def formatSheet():
    # autoFit()
    # time.sleep(5)

    wb = load_workbook('res.xlsx')



    changeDim(wb, wrapRows)
    set_border(wb, allCells)
    formatRatingRow(wb, ratingRows, localeRows, assignmentRows)
    formatGrayColumn(wb, targetRows, grayColumn)
    mainSheet = wb.get_sheet_by_name("Sheet1")

    for i in ['G' + str(j) for j in range(23, 42)]:
        mainSheet[i].alignment = alignmentR
        # wb.save("res.xlsx")
    for i in ['G' + str(j) for j in range(11, 15)]:
        mainSheet[i].alignment = alignmentL

    wb.save("res.xlsx")



    print "Formatting done"



# formatSheet()
