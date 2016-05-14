__author__ = 'rjain1'
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from testing import returnDatafromTemplate


def createSingleRow(indexAlpha, num):
    indexAlpha = [i + str(num) for i in indexAlpha]
    return indexAlpha


def createRows():
    rows = []

    indexAlpha = ['A', 'B', 'C', 'D', 'E', 'F']

    for i in range(1, 23):
        rows.append(createSingleRow(indexAlpha, i))
    return rows[1:]


def prepareSheet(wb):
    sheets = wb.get_sheet_names()
    print sheets
    if "Sheet" not in sheets:
        wb.create_sheet(title="Sheet1")

    sheets.remove("ICICI_Sites")
    sheets.remove("Ireland")
    sheets.remove("RBC")

    wb.save("stats.xlsx")

    return sheets


def mapValues(wb, index, values):
    mainSheet = wb.get_sheet_by_name("Sheet1")

    data = zip(index, values)
    for i in data:
        mainSheet[i[0]] = i[1]
    wb.save("stats.xlsx")


def processSheet():
    wb = openpyxl.load_workbook('stats.xlsx', data_only=True)
    temp = openpyxl.load_workbook('stats.xlsx', data_only=True)

    print prepareSheet(wb)

    mainSheet = wb.get_sheet_by_name("Sheet1")
    mainSheet["A1"] = "test"
    wb.save("stats.xlsx")

    rowsIndex = createRows()
    print rowsIndex
    print len(rowsIndex)
    x = returnDatafromTemplate()
    print len(x)
    print rowsIndex
    print x
    # for i in x:
    #     print i
    for i in range(len(x)):
        mapValues(wb, rowsIndex[i], x[i])
    # mapValues(wb, rowsIndex[0], x[0])
    print "Added blueprint"

# processSheet()
